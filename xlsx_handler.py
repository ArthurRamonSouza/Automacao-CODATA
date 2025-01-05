import openpyxl
from typing import List, Any


class XlsxHandler:
  """
        XlsxHandler - A utility class for handling Excel (.xlsx) files using the openpyxl library.

        Methods:
        - open_workbook(file_path): Opens an Excel workbook from the specified file path.
        - get_sheet(workbook, sheet_name): Retrieves a specific sheet from the given workbook.
        - close_workbook(workbook, file_name): Saves and closes the workbook, printing a message.
        - edit_sheet_row(sheet, row_number, columns, new_values): Edits a row in the specified sheet with new values.

        Usage:
        handler = XlsxHandler()
        workbook = handler.open_workbook("example.xlsx")
        sheet = handler.get_sheet(workbook, "Sheet1")

        # Perform operations on the workbook or sheet

        handler.edit_sheet_row(sheet, 2, ['A', 'B', 'C'], [1, 2, 3])
        handler.close_workbook(workbook, "modified_example.xlsx")
        """

  @staticmethod
  def open_workbook(file_path: str):
    """
                Opens an Excel workbook from the specified file path.

                Parameters:
                - file_path (str): The path to the Excel file.

                Returns:
                - Workbook (openpyxl.Workbook): The opened workbook object.
                """
    return openpyxl.load_workbook(file_path)

  @staticmethod
  def get_sheet(workbook: openpyxl.Workbook,
                sheet_name: str) -> openpyxl.worksheet.worksheet:
    """
            Retrieves a specific sheet from the given workbook.

            Parameters:
            - workbook (openpyxl.Workbook): The Excel workbook.
            - sheet_name (str): The name of the sheet to retrieve.

            Returns:
            - Worksheet (openpyxl.worksheet.worksheet.Worksheet): The specified sheet object.
            """
    return workbook[sheet_name]

  @staticmethod
  def close_workbook(workbook: openpyxl.Workbook, file_name: str) -> None:
    """
            Saves and closes the workbook, printing a closing message.

            Parameters:
            - workbook (openpyxl.Workbook): The Excel workbook.
            - file_name (str): The name to save the workbook as.
            """
    workbook.save(file_name)
    workbook.close()
    print("Workbook closed!")

  @staticmethod
  def edit_sheet_row(sheet: openpyxl.worksheet.worksheet, row_number: int,
                     columns: List[str], new_values: List[Any]) -> None:
    """
            Edits a row in the specified sheet with new values.

            Parameters:
            - sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel sheet to edit.
            - row_number (int): The row number to edit.
            - columns (List[str]): List of column letters to edit.
            - new_values (List[Any]): List of new values to replace in the specified columns.
            """
    try:
      if XlsxHandler.is_valid_row(row_number, sheet):
        for col_idx, col_letter in enumerate(columns):
          cell_address = f"{col_letter}{row_number}"
          sheet[cell_address] = new_values[col_idx]
      else:
        print(f"Invalid row number: {row_number}")

    except Exception as e:
      print(f"An error occurred on XlsxHandler class: {e}")

  @staticmethod
  def is_valid_row(row_number: int,
                   sheet: openpyxl.worksheet.worksheet) -> bool:
    """
            Checks if the given row number is valid for the provided sheet.

            Parameters:
            - row_number (int): The row number to check.
            - sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel sheet to validate against.

            Returns:
            - bool: True if the row number is valid, False otherwise.
            """
    if not 1 <= row_number <= sheet.max_row:
      print(
          f"Error: Invalid row number {row_number}. The file has {sheet.max_row} rows."
      )
      return False
    return True
